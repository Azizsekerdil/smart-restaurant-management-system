"""Dışa aktarma: CSV, Excel (openpyxl) ve PDF (ReportLab).

WeasyPrint yerine ReportLab tercih edildi: Windows'ta GTK bağımlılığı
gerektirmez, tek `pip install` ile çalışır.
"""

from __future__ import annotations

import csv
import io
import logging
import os
from datetime import date
from decimal import Decimal
from functools import lru_cache
from pathlib import Path

from django.conf import settings
from django.http import HttpResponse
from django.utils import timezone
from django.utils.functional import Promise
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    TableStyle,
)
from reportlab.platypus import (
    Table as PdfTable,
)

from apps.core.utils import format_money

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)


# ------------------------------------------------------------------
#  PDF yazı tipi
# ------------------------------------------------------------------
#: ReportLab'ın yerleşik Helvetica'sı WinAnsi kodlamasını kullanır ve
#: **Türkçe'ye özgü harfleri içermez**: ş, ğ, ı ve İ çıktıda siyah kutu
#: olarak görünür ("Satış Raporu" -> "Sat■■ Raporu"). ç, ö, ü sorunsuzdur
#: çünkü WinAnsi tablosunda yer alırlar — bu yüzden hata gözden kaçabilir.
#: Bu yüzden sistemden Türkçe destekli bir TrueType yazı tipi yüklenir.
_FONT_CANDIDATES = (
    ("SegoeUI", "segoeui.ttf", "segoeuib.ttf"),
    ("Arial", "arial.ttf", "arialbd.ttf"),
    ("Calibri", "calibri.ttf", "calibrib.ttf"),
    ("DejaVuSans", "DejaVuSans.ttf", "DejaVuSans-Bold.ttf"),
)


@lru_cache(maxsize=1)
def pdf_fonts() -> tuple[str, str]:
    """(normal, kalın) yazı tipi adlarını döndürür; bir kez kaydeder."""
    search_dirs = [
        Path(os.environ.get("WINDIR", "C:/Windows")) / "Fonts",
        Path("/usr/share/fonts/truetype/dejavu"),
        Path("/usr/share/fonts/truetype/msttcorefonts"),
    ]
    for name, regular, bold in _FONT_CANDIDATES:
        for directory in search_dirs:
            regular_path, bold_path = directory / regular, directory / bold
            if regular_path.is_file() and bold_path.is_file():
                try:
                    pdfmetrics.registerFont(TTFont(name, str(regular_path)))
                    pdfmetrics.registerFont(TTFont(f"{name}-Bold", str(bold_path)))
                except Exception:  # pragma: no cover - bozuk yazı tipi dosyası
                    # Bozuk bir yazı tipi yüzünden rapor üretimi durmamalı;
                    # sıradaki adaya geçilir.
                    logger.warning("Yazı tipi yüklenemedi: %s", regular_path, exc_info=True)
                    continue
                return name, f"{name}-Bold"

    # Son çare: Türkçe harfler kaybolur ama belge yine de üretilir.
    logger.warning(
        "Türkçe destekli PDF yazı tipi bulunamadı; ş/ğ/ı/İ harfleri çıktıda görünmeyecek."
    )
    return "Helvetica", "Helvetica-Bold"


def _cell_value(value):
    """Hücre değerini openpyxl'in kabul ettiği türe indirger.

    Çevrilebilir etiketler ``gettext_lazy`` ile sarılıdır ve gerçek bir
    ``str`` değildir; openpyxl bunları "Excel'e çevrilemez" diyerek
    reddeder. Aynı şekilde Decimal de float'a indirgenir.
    """
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, Promise):  # gettext_lazy sonucu
        return str(value)
    return value


# ------------------------------------------------------------------
#  CSV
# ------------------------------------------------------------------
def csv_response(filename: str, headers: list[str], rows: list[list]) -> HttpResponse:
    response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
    response["Content-Disposition"] = f'attachment; filename="{filename}.csv"'
    # UTF-8 BOM: Excel'in Türkçe karakterleri doğru göstermesi için.
    response.write("﻿")
    writer = csv.writer(response, delimiter=";")
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return response


# ------------------------------------------------------------------
#  Excel
# ------------------------------------------------------------------
def excel_response(
    filename: str,
    sheets: dict[str, tuple[list[str], list[list]]],
    *,
    title: str = "",
) -> HttpResponse:
    """Birden çok sayfalı Excel dosyası üretir.

    `sheets` = {"Sayfa adı": (başlıklar, satırlar)}
    """
    workbook = Workbook()
    workbook.remove(workbook.active)

    for sheet_name, (headers, rows) in sheets.items():
        sheet = workbook.create_sheet(title=sheet_name[:31])
        if title:
            sheet.append([title])
            sheet.append([f"Oluşturulma: {timezone.localtime():%d.%m.%Y %H:%M}"])
            sheet.append([])
            sheet["A1"].font = Font(bold=True, size=14)

        header_row = sheet.max_row + 1
        sheet.append(headers)
        for cell in sheet[header_row]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in rows:
            sheet.append([_cell_value(value) for value in row])

        for index, header in enumerate(headers, start=1):
            width = max(len(str(header)) + 4, 14)
            for row in rows[:200]:
                if index <= len(row):
                    width = max(width, min(len(str(row[index - 1])) + 3, 50))
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)

    buffer = io.BytesIO()
    workbook.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
    return response


# ------------------------------------------------------------------
#  PDF
# ------------------------------------------------------------------
def _styles():
    base = getSampleStyleSheet()
    regular, bold = pdf_fonts()

    # Hazır stiller Helvetica'ya bağlıdır; Türkçe harfler için hepsini
    # kayıtlı yazı tipine çeviriyoruz.
    for name in ("Normal", "Title", "Heading2"):
        base[name].fontName = bold if name != "Normal" else regular

    return {
        "title": ParagraphStyle(
            "RTitle",
            parent=base["Title"],
            fontName=bold,
            fontSize=16,
            spaceAfter=6,
            textColor=colors.HexColor("#1F3A5F"),
        ),
        "sub": ParagraphStyle(
            "RSub",
            parent=base["Normal"],
            fontName=regular,
            fontSize=9,
            textColor=colors.grey,
            spaceAfter=12,
        ),
        "h2": ParagraphStyle(
            "RH2",
            parent=base["Heading2"],
            fontName=bold,
            fontSize=12,
            spaceBefore=10,
            spaceAfter=4,
        ),
        "normal": ParagraphStyle("RNormal", parent=base["Normal"], fontName=regular),
        "right": ParagraphStyle(
            "RRight", parent=base["Normal"], fontName=regular, alignment=TA_RIGHT
        ),
        "center": ParagraphStyle(
            "RCenter", parent=base["Normal"], fontName=regular, alignment=TA_CENTER
        ),
        "small": ParagraphStyle(
            "RSmall", parent=base["Normal"], fontName=regular, fontSize=8, textColor=colors.grey
        ),
        "bold": ParagraphStyle("RBold", parent=base["Normal"], fontName=bold),
    }


def _pdf_table(headers: list[str], rows: list[list], *, col_widths=None) -> PdfTable:
    regular, bold = pdf_fonts()
    data = [headers] + [[str(c) for c in row] for row in rows]
    table = PdfTable(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3A5F")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, -1), regular),
                ("FONTNAME", (0, 0), (-1, 0), bold),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCCCCC")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return table


def pdf_report(
    title: str,
    sections: list[dict],
    *,
    subtitle: str = "",
    footer_note: str = "",
) -> bytes:
    """Genel amaçlı PDF rapor üreticisi.

    `sections` = [{"heading": str, "headers": [...], "rows": [[...]], "text": str}]
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
        title=title,
        author=settings.RESTAURANT["NAME"],
    )
    styles = _styles()
    story = [
        Paragraph(_esc(settings.RESTAURANT["NAME"]), styles["title"]),
        Paragraph(_esc(title), styles["h2"]),
    ]
    if subtitle:
        story.append(Paragraph(_esc(subtitle), styles["sub"]))
    story.append(Paragraph(f"Oluşturulma: {timezone.localtime():%d.%m.%Y %H:%M}", styles["small"]))
    story.append(Spacer(1, 8))

    for section in sections:
        if section.get("heading"):
            story.append(Paragraph(_esc(section["heading"]), styles["h2"]))
        if section.get("text"):
            story.append(Paragraph(_esc(section["text"]), styles["normal"]))
            story.append(Spacer(1, 4))
        if section.get("headers"):
            story.append(_pdf_table(section["headers"], section.get("rows", [])))
            story.append(Spacer(1, 10))

    if footer_note:
        story.append(Spacer(1, 12))
        story.append(Paragraph(_esc(footer_note), styles["small"]))

    doc.build(story)
    return buffer.getvalue()


def _esc(text) -> str:
    """ReportLab Paragraph için XML kaçışı."""
    return str(text).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def pdf_response(filename: str, content: bytes) -> HttpResponse:
    response = HttpResponse(content, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}.pdf"'
    return response


# ------------------------------------------------------------------
#  Özel belgeler
# ------------------------------------------------------------------
def order_receipt_pdf(order) -> bytes:
    """80 mm termal fiş genişliğinde adisyon PDF'i."""
    buffer = io.BytesIO()
    width = 80 * mm
    height = (120 + len(list(order.active_items)) * 12) * mm
    doc = SimpleDocTemplate(
        buffer,
        pagesize=(width, height),
        leftMargin=4 * mm,
        rightMargin=4 * mm,
        topMargin=4 * mm,
        bottomMargin=4 * mm,
    )
    styles = _styles()
    small = ParagraphStyle("Tiny", parent=styles["normal"], fontSize=7, leading=9)
    center = ParagraphStyle("TinyCenter", parent=small, alignment=TA_CENTER)

    story = [
        Paragraph(f"<b>{_esc(settings.RESTAURANT['NAME'])}</b>", center),
        Paragraph(f"Adisyon: {_esc(order.number)}", small),
        Paragraph(f"Tarih: {timezone.localtime(order.opened_at):%d.%m.%Y %H:%M}", small),
        Paragraph(
            f"Masa: {_esc(order.table.name if order.table_id else order.get_order_type_display())}",
            small,
        ),
        Spacer(1, 4),
    ]

    rows = []
    for item in order.active_items:
        rows.append(
            [
                f"{item.quantity:g} x {item.product_name[:22]}",
                format_money(item.net_total),
            ]
        )
    regular, bold = pdf_fonts()
    table = PdfTable(rows, colWidths=[46 * mm, 24 * mm])
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), regular),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("LINEBELOW", (0, 0), (-1, -1), 0.2, colors.HexColor("#DDDDDD")),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]
        )
    )
    story.append(table)
    story.append(Spacer(1, 4))

    summary = [
        ["Ara toplam", format_money(order.subtotal)],
    ]
    if order.order_discount_total:
        summary.append(["İndirim", f"-{format_money(order.order_discount_total)}"])
    if order.service_charge:
        summary.append(["Servis", format_money(order.service_charge)])
    summary.append(["KDV (dahil)", format_money(order.tax_total)])
    summary.append(["TOPLAM", format_money(order.grand_total)])

    summary_table = PdfTable(summary, colWidths=[46 * mm, 24 * mm])
    summary_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), regular),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("FONTNAME", (0, -1), (-1, -1), bold),
                ("LINEABOVE", (0, -1), (-1, -1), 0.5, colors.black),
            ]
        )
    )
    story.append(summary_table)
    story.append(Spacer(1, 6))
    story.append(Paragraph("Bizi tercih ettiğiniz için teşekkürler!", center))
    story.append(
        Paragraph(
            "Bu belge mali değeri olmayan bir bilgi fişidir.",
            ParagraphStyle("Note", parent=center, fontSize=6, textColor=colors.grey),
        )
    )
    doc.build(story)
    return buffer.getvalue()


def daily_closing_pdf(closing) -> bytes:
    """Gün sonu (Z benzeri) kapanış raporu."""
    sections = [
        {
            "heading": "Satış Özeti",
            "headers": ["Kalem", "Değer"],
            "rows": [
                ["Sipariş sayısı", closing.order_count],
                ["Misafir sayısı", closing.guest_count],
                ["Brüt satış", format_money(closing.gross_sales)],
                ["İndirimler", format_money(closing.discount_total)],
                ["İadeler", format_money(closing.refund_total)],
                ["Servis bedeli", format_money(closing.service_charge_total)],
                ["KDV (dahil)", format_money(closing.tax_total)],
                ["Net satış", format_money(closing.net_sales)],
                ["Bahşiş", format_money(closing.tip_total)],
                ["Ortalama adisyon", format_money(closing.average_ticket)],
                ["Kişi başı ortalama", format_money(closing.average_per_guest)],
                [
                    "İptal edilen adisyon",
                    f"{closing.void_count} ({format_money(closing.void_total)})",
                ],
            ],
        },
        {
            "heading": "Ödeme Dağılımı",
            "headers": ["Yöntem", "Tutar"],
            "rows": [[k, v] for k, v in (closing.payment_breakdown or {}).items()] or [["-", "0"]],
        },
        {
            "heading": "Kategori Dağılımı",
            "headers": ["Kategori", "Ciro"],
            "rows": [[k, v] for k, v in (closing.category_breakdown or {}).items()] or [["-", "0"]],
        },
        {
            "heading": "Kasa",
            "headers": ["Kalem", "Tutar"],
            "rows": [
                ["Beklenen nakit", format_money(closing.cash_expected)],
                ["Sayılan nakit", format_money(closing.cash_counted)],
                ["Fark", format_money(closing.cash_variance)],
            ],
        },
    ]
    return pdf_report(
        f"Gün Sonu Raporu — {closing.closing_date:%d.%m.%Y}",
        sections,
        subtitle=f"Kapatan: {closing.closed_by.display_name if closing.closed_by_id else '—'}",
        footer_note=closing.legal_notice,
    )


def sales_report_excel(start: date, end: date) -> HttpResponse:
    """Kapsamlı satış raporu (çok sayfalı Excel)."""
    from apps.reports import services

    summary = services.dashboard_metrics(end)
    products = services.top_products(start, end, limit=200)
    categories = services.category_breakdown(start, end)
    payments = services.payment_breakdown(start, end)
    staff = services.staff_sales_report(start, end)
    profit = services.profitability_report(start, end, limit=200)

    sheets = {
        "Özet": (
            ["Kalem", "Değer"],
            [
                ["Dönem", f"{start:%d.%m.%Y} - {end:%d.%m.%Y}"],
                ["Ciro (son gün)", summary["revenue"]],
                ["Sipariş sayısı", summary["order_count"]],
                ["Ortalama adisyon", summary["average_ticket"]],
                ["İptal oranı (%)", summary["cancel_rate"]],
                ["Doluluk oranı (%)", summary["occupancy_rate"]],
            ],
        ),
        "Ürünler": (
            ["Ürün", "Kategori", "Adet", "Ciro"],
            [
                [
                    p["product__name"],
                    p["product__category__name"] or "-",
                    p["total_quantity"],
                    p["revenue"],
                ]
                for p in products
            ],
        ),
        "Kategoriler": (
            ["Kategori", "Ciro", "Adet", "Pay (%)"],
            [[c["category"], c["revenue"], c["quantity"], c["percent"]] for c in categories],
        ),
        "Ödemeler": (
            ["Yöntem", "Tutar", "İşlem", "Pay (%)"],
            [[p["label"], p["total"], p["count"], p["percent"]] for p in payments],
        ),
        "Personel": (
            ["Personel", "Ciro", "Sipariş", "Misafir", "Ort. Adisyon", "İndirim"],
            [
                [
                    s["name"],
                    s["revenue"],
                    s["orders"],
                    s["guests"],
                    s["average_ticket"],
                    s["discounts"],
                ]
                for s in staff
            ],
        ),
        "Karlilik": (
            ["Ürün", "Adet", "Ciro", "Birim Maliyet", "Toplam Maliyet", "Kâr", "Marj (%)"],
            [
                [
                    r["product"].name,
                    r["quantity"],
                    r["revenue"],
                    r["unit_cost"],
                    r["total_cost"],
                    r["profit"],
                    r["margin_percent"],
                ]
                for r in profit
            ],
        ),
    }
    return excel_response(
        f"satis-raporu-{start:%Y%m%d}-{end:%Y%m%d}",
        sheets,
        title=f"Satış Raporu — {settings.RESTAURANT['NAME']}",
    )


def statistics_workbook(data: dict) -> HttpResponse:
    """İstatistik merkezinin tüm bölümlerini çok sayfalı Excel olarak verir."""
    start, end = data["start"], data["end"]
    comparison = data["comparison"]
    matrix = data["matrix"]

    sheets = {
        "Karsilastirma": (
            ["Ölçüt", "Bu dönem", "Önceki dönem", "Değişim (%)"],
            [[m["label"], m["current"], m["previous"], m["change"]] for m in comparison["metrics"]],
        ),
        "Gunluk": (
            ["Gün", "Ciro"],
            list(zip(data["daily"]["labels"], data["daily"]["current"], strict=False)),
        ),
        "Haftanin gunleri": (
            ["Gün", "Toplam ciro", "Sipariş", "Gözlenen gün", "Günlük ortalama"],
            [
                [w["name"], w["revenue"], w["orders"], w["observed_days"], w["average_revenue"]]
                for w in data["weekday"]
            ],
        ),
        "Gun x Saat": (
            ["Gün", *[f"{hour}:00" for hour in matrix["hours"]]],
            [[row["name"], *[cell["average"] for cell in row["cells"]]] for row in matrix["rows"]],
        ),
        "Kategoriler": (
            ["Kategori", "Ciro", "Adet", "Pay (%)"],
            [
                [c["category"], c["revenue"], c["quantity"], c["percent"]]
                for c in data["categories"]
            ],
        ),
        "Odemeler": (
            ["Yöntem", "Tutar", "İşlem", "Pay (%)"],
            [[p["label"], p["total"], p["count"], p["percent"]] for p in data["payments"]],
        ),
        "Siparis turleri": (
            ["Tür", "Tutar", "Adet"],
            [[o["label"], o["total"], o["count"]] for o in data["order_types"]],
        ),
        "Personel": (
            ["Personel", "Ciro", "Sipariş", "Misafir", "Ort. Adisyon"],
            [
                [s["name"], s["revenue"], s["orders"], s["guests"], s["average_ticket"]]
                for s in data["staff"]
            ],
        ),
        "Urunler": (
            ["Ürün", "Kategori", "Adet", "Ciro"],
            [
                [
                    p["product__name"],
                    p["product__category__name"] or "-",
                    p["total_quantity"],
                    p["revenue"],
                ]
                for p in data["top_products"]
            ],
        ),
        "Musteriler": (
            ["Müşteri", "Ciro", "Ziyaret", "Ortalama"],
            [
                [
                    str(row["customer"]) if row["customer"] else "-",
                    row["revenue"],
                    row["visits"],
                    row["average"],
                ]
                for row in data["customers"]["top_customers"]
            ],
        ),
        "Stok": (
            ["Kalem", "Değer"],
            [
                ["Fire tutarı", data["inventory"]["waste_value"]],
                ["Satıştan tüketim", data["inventory"]["consumption_value"]],
                ["Satın alma", data["inventory"]["purchase_value"]],
                ["Fire oranı (%)", data["inventory"]["waste_percent"]],
                ["Mevcut stok değeri", data["inventory"]["stock_value"]],
                ["Kritik seviyedeki malzeme", data["inventory"]["critical_count"]],
            ],
        ),
    }
    return excel_response(
        f"istatistik-{start:%Y%m%d}-{end:%Y%m%d}",
        sheets,
        title=f"İstatistik Merkezi — {settings.RESTAURANT['NAME']} ({start:%d.%m.%Y} – {end:%d.%m.%Y})",
    )
